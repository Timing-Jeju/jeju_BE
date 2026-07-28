from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE = Path("outputs/timing_jeju_requirements_fixed.xlsx")
OUT_REPO = Path("outputs/timing_jeju_requirements_ordered.xlsx")
OUT_DOWNLOADS = Path("/Users/josephuk77/Downloads/timing_jeju_requirements_ordered.xlsx")


ORDER = [
    # 0. 먼저 합의해야 하는 계약/데모 기준
    "NFR-005", "NFR-004", "DATA-014", "NFR-006",

    # 1. 외부 API와 데이터 기반
    "DATA-001", "DATA-002", "DATA-005", "DATA-006", "DATA-007",
    "DATA-012", "DATA-013",

    # 2. AI 입력/가드레일
    "AI-001", "AI-002", "FR-002",

    # 3. 사용자 입력과 관광지 선택 흐름
    "FR-001", "FR-031", "FR-032", "FR-033", "UI-013",
    "FR-003", "FR-004", "FR-005", "FR-006",
    "FR-007", "FR-008", "UI-001", "UI-002",

    # 4. 핵심 일정 계산 엔진
    "FR-009", "FR-010", "FR-034", "FR-013", "FR-014", "FR-015",
    "FR-017", "FR-018", "FR-011", "FR-012", "FR-025",

    # 5. 결과/타임라인/지도 딥링크 화면
    "UI-003", "UI-004", "UI-005", "UI-006", "DATA-010",
    "UI-007", "FR-016", "DATA-011", "FR-022", "UI-011",

    # 6. 복구안과 설명
    "FR-019", "FR-020", "FR-021", "UI-008", "AI-003", "AI-004", "UI-010",

    # 7. 대표 코스 제한 추천
    "FR-023", "DATA-003", "FR-024", "UI-009",

    # 8. 안정성/성능 마감
    "NFR-001", "NFR-002", "NFR-003",

    # 9. Phase 2 확장
    "DATA-004", "DATA-008", "DATA-009", "FR-026", "FR-027",
    "FR-028", "FR-029", "UI-012",

    # 10. Phase 3
    "FR-030",
]


WORK_MEMO = {
    "NFR-005": "1. API/DTO 계약 먼저 고정",
    "NFR-004": "1. 프론트 fixture 화면 병렬 개발 준비",
    "DATA-014": "1. fixture와 실제 데이터가 같은 엔진을 타도록 고정",
    "NFR-006": "1. 본선 대표 시나리오 검수 기준 고정",
    "DATA-001": "2. TourAPI 장소 검색 기반",
    "DATA-002": "2. 관광지 좌표/상세 정보 기반",
    "DATA-005": "2. 정류장 후보 검색 기반",
    "DATA-006": "2. 노선 후보 검색 기반",
    "DATA-007": "2. 시간표 계산 기반",
    "DATA-012": "2. 데이터 신뢰 필드 계약",
    "DATA-013": "2. 데이터 헬스 확인",
    "AI-001": "3. 자연어 입력 구조화",
    "AI-002": "3. AI 교통 사실 생성 금지",
    "FR-002": "3. 자연어 입력 기능 연결",
    "FR-001": "4. 기본 입력 폼",
    "FR-031": "4. 자연어 없이 직접 일정 만들기",
    "FR-032": "4. 직접 일정 장소 추가/삭제",
    "FR-033": "4. 직접 일정 순서 조정",
    "UI-013": "4. 직접 일정 빌더 화면",
    "FR-003": "4. 희망 관광지 입력",
    "FR-004": "4. 관광지 후보 조회",
    "FR-005": "4. 관광지 후보 선택",
    "FR-006": "4. 후보 카드 정보",
    "FR-007": "4. 체류시간 수정",
    "FR-008": "4. 필수 방문 장소",
    "UI-001": "4. 입력 화면",
    "UI-002": "4. 후보 화면",
    "FR-009": "5. 대표 코스 계산",
    "FR-010": "5. 일정안 생성",
    "FR-034": "5. 직접 작성 일정 실행 가능성 검사",
    "FR-013": "5. 정류장 후보 매칭",
    "FR-014": "5. 구간별 버스 정보",
    "FR-015": "5. 출발 권장 시각",
    "FR-017": "5. 안전도 점수",
    "FR-018": "5. 위험 구간",
    "FR-011": "5. 일정안 비교",
    "FR-012": "5. 최종 일정 선택",
    "FR-025": "5. 일정 변경 시 재계산",
    "UI-003": "6. 일정안 선택 화면",
    "UI-004": "6. 결과 요약",
    "UI-005": "6. 타임라인",
    "UI-006": "6. 버스 정보 카드",
    "DATA-010": "6. 지도 딥링크",
    "UI-007": "6. 지도 위치 확인",
    "FR-016": "6. 현재 상태 안내",
    "DATA-011": "6. 위치 정보/수동 대체",
    "FR-022": "6. 위치 권한 대체 흐름",
    "UI-011": "6. 데이터 기준 확인 화면",
    "FR-019": "7. 버스 놓침 영향",
    "FR-020": "7. 놓침 복구안",
    "FR-021": "7. 막차 위험",
    "UI-008": "7. 복구 화면",
    "AI-003": "7. 계산 결과 설명",
    "AI-004": "7. 직접 일정 실행 가능성 설명",
    "UI-010": "7. UX 라이팅",
    "FR-023": "8. 여유 시간 감지",
    "DATA-003": "8. 대표 코스 주변 장소 조회",
    "FR-024": "8. 주변 안전 추천",
    "UI-009": "8. 추천 화면",
    "NFR-001": "9. 응답 시간",
    "NFR-002": "9. 장애 대응",
    "NFR-003": "9. 위치정보 최소 수집",
    "DATA-004": "10. 행사/축제 확장",
    "DATA-008": "10. 실시간 도착 확장",
    "DATA-009": "10. 실시간 위치 확장",
    "FR-026": "10. 일정 저장/비교 확장",
    "FR-027": "10. 짐 보관 확장",
    "FR-028": "10. 날씨 보정 확장",
    "FR-029": "10. 택시 플랜B 확장",
    "UI-012": "10. RTO 발표 보조",
    "FR-030": "11. 네이티브 위젯 확장",
}


def thin_border():
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def main():
    wb = load_workbook(SOURCE)
    ws = wb["요구사항"]

    header = [cell.value for cell in ws[1]]
    rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    by_id = {row[1]: row for row in rows}

    missing = [rid for rid in ORDER if rid not in by_id]
    extra = [row[1] for row in rows if row[1] not in ORDER]
    if missing or extra:
        raise RuntimeError(f"missing={missing}, extra={extra}")

    ordered_rows = [by_id[rid] for rid in ORDER]
    for index, row in enumerate(ordered_rows, 1):
        row[0] = index
        memo = row[13] or ""
        work_memo = WORK_MEMO.get(row[1], "")
        row[13] = f"{work_memo} / {memo}" if memo else work_memo

    ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for row in ordered_rows:
        ws.append(row)

    widths = [6, 12, 12, 16, 24, 58, 34, 42, 10, 10, 12, 12, 8, 28]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    phase_colors = {
        "MVP": "EAF3F8",
        "Phase 2": "FFF2CC",
        "Phase 3": "E7E6E6",
    }
    for row in ws.iter_rows(min_row=2):
        phase = row[10].value
        fill = PatternFill("solid", fgColor=phase_colors.get(phase, "FFFFFF"))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            cell.fill = fill
        ws.row_dimensions[row[0].row].height = 66

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"

    summary = wb["요약"]
    summary["A15"] = "정렬 기준"
    summary["B15"] = "개발 실행 순서 기준: 계약/fixture → 데이터 → AI 입력 → 입력 UI → 계산 엔진 → 결과 화면 → 복구 → 추천 → 안정화 → 확장"
    summary["A15"].font = Font(bold=True)
    summary["A15"].border = thin_border()
    summary["B15"].border = thin_border()
    summary["B15"].alignment = Alignment(wrap_text=True, vertical="top")

    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_REPO)
    print(OUT_REPO.resolve())


if __name__ == "__main__":
    main()
