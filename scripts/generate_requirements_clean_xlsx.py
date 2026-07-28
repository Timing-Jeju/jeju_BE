from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from copy import deepcopy

from generate_requirements_like_template import requirements as base_requirements


OUT_REPO = Path("outputs/timing_jeju_requirements_fixed.xlsx")
OUT_DOWNLOADS = Path("/Users/josephuk77/Downloads/timing_jeju_requirements_fixed.xlsx")

HEADERS = [
    "No", "ID", "구분", "모듈", "요구사항명", "상세 요구사항", "연동/하위 기능",
    "검수 기준", "필요성", "우선순위", "단계", "상태", "체크", "비고"
]


def build_requirements():
    reqs = deepcopy(base_requirements)

    for req in reqs:
        req_id = req[0]

        if req_id == "FR-013":
            req[4] = (
                "시스템은 관광지 주변의 탑승 가능 정류장과 하차 후 접근 가능한 정류장 후보를 찾아야 한다. "
                "MVP에서는 정류장을 확정값처럼 표현하지 않고, 가까운 정류장 후보와 지도 확인이 필요한 상태로 안내해야 한다."
            )
            req[5] = "정류장 좌표, 관광지 좌표, 반경 검색, 노선 방향 판단, 후보형 표시"
            req[6] = "각 이동 구간에 탑승/하차 정류장 후보가 표시되고, 방향이 불확실한 경우 후보 또는 확인 필요 상태로 표시된다."

        if req_id == "FR-014":
            req[4] = (
                "관광지 간 이동 구간마다 탑승 정류장 후보, 하차 정류장 후보, 버스 노선, 출발 시각, "
                "도착 예상 시각을 제공해야 한다. 시간표 데이터가 주요 지점 기준인 경우 이를 함께 표시해야 한다."
            )
            req[5] = "정류장 후보 매칭, 버스 노선 조회, 시간표 조회, 구간 상세 카드, 데이터 기준 문구"
            req[6] = "각 이동 구간별 버스 정보와 데이터 기준이 함께 표시된다."

        if req_id == "FR-024":
            req[3] = "대표 코스 주변 안전 추천"
            req[4] = (
                "MVP에서는 제주 동쪽 대표 코스 주변으로 범위를 제한해, 남는 시간 안에 실제로 가능한 TourAPI 후보를 2~3개 추천해야 한다. "
                "제주 전역 추천이나 복잡한 개인화 추천은 Phase 2로 미룬다."
            )
            req[5] = "TourAPI 위치기반 조회, 대표 코스 주변 후보 seed/cache, 거리 계산, 다음 버스 복귀 버퍼"
            req[6] = "대표 시나리오에서 남는 시간이 생기면 다음 버스에 늦지 않는 후보만 가능/주의 상태로 표시된다."
            req[7] = "필수"
            req[8] = "상"
            req[9] = "MVP"
            req[12] = "CEO/제안서 반영"

        if req_id == "DATA-003":
            req[3] = "대표 코스 주변 장소 조회"
            req[4] = (
                "현재 위치, 관광지, 정류장 주변의 관광지, 음식점, 카페 후보를 조회할 수 있어야 한다. "
                "MVP에서는 대표 코스 주변 후보를 seed/cache와 TourAPI 위치기반 조회로 제한해 사용한다."
            )
            req[5] = "TourAPI 위치기반 관광정보, 반경 검색, 카테고리 필터, 대표 코스 cache"
            req[6] = "대표 코스 주변에서 남는 시간 추천 후보가 조회되고, 데이터가 없으면 추천 없음 상태가 표시된다."
            req[7] = "필수"
            req[8] = "상"
            req[9] = "MVP"
            req[12] = "제안서/CEO 계획 반영"

        if req_id == "DATA-005":
            req[4] = (
                "관광지 주변 정류소 탐색을 위해 제주 버스 정류장 ID, 정류장명, 위도, 경도를 DB에 저장해야 한다. "
                "정류장 데이터는 확정 승차 위치가 아니라 가까운 정류장 후보 산출에 사용한다."
            )
            req[6] = "DB에서 관광지 좌표 기준 가까운 정류장 후보를 거리순으로 조회할 수 있고, 후보형 문구로 표시된다."

        if req_id == "DATA-007":
            req[4] = (
                "버스 출발 시각, 대기시간, 막차 위험 판단을 위해 제주 버스 시간표 데이터를 활용해야 한다. "
                "MVP에서는 동쪽 코스에 필요한 101/201 중심 노선과 주요 시간표 지점 기준으로 우선 등록한다."
            )
            req[5] = "노선별 시간표, 요일별 시간표, 막차 정보, XLSX hash, 시간표 기준일, 시간표 DB"
            req[6] = "대표 코스의 구간별 출발 시각과 대기시간이 시간표 기준으로 계산되고, 공식 시간표 주요 지점 기준임이 표시된다."

        if req_id == "DATA-010":
            req[3] = "지도 딥링크 우선 연동"
            req[4] = (
                "MVP에서는 네이티브 지도 SDK에 의존하지 않고, 관광지와 정류장 위치를 카카오맵/네이버지도 딥링크로 열 수 있어야 한다. "
                "앱 내부 지도나 WebView 지도 미리보기는 선택 기능으로 둔다."
            )
            req[5] = "Kakao/Naver deep link, OpenMapButton, 선택적 WebView preview"
            req[6] = "각 장소와 정류장 후보에서 외부 지도 열기 버튼이 동작하고, 지도 SDK가 없어도 핵심 일정 결과를 볼 수 있다."
            req[12] = "엔지니어링 리뷰 반영"

        if req_id == "DATA-012":
            req[3] = "데이터 신뢰 정보 표시"
            req[4] = (
                "일정 결과와 추천 결과에는 TourAPI, 버스 시간표, fixture 여부, fallback 여부 등 어떤 데이터를 기준으로 계산했는지 표시해야 한다. "
                "계산 결과에는 source, dataVersion, confidence, stale, fallbackUsed 같은 신뢰 필드를 포함해야 한다."
            )
            req[5] = "source, dataVersion, confidence, stale, fallbackUsed, generatedAt, warnings"
            req[6] = "결과 화면과 API 응답에서 데이터 기준, 신뢰도, fallback 여부를 확인할 수 있다."
            req[12] = "CEO/Eng guardrail 반영"

        if req_id == "UI-007":
            req[3] = "지도 위치 확인 화면"
            req[4] = (
                "관광지, 탑승 정류장 후보, 하차 정류장 후보의 위치를 확인할 수 있어야 한다. "
                "MVP에서는 앱 내부 통합 지도보다 외부 지도 링크와 간단한 위치 확인 UI를 우선한다."
            )
            req[5] = "장소/정류장 카드, 지도 열기 버튼, 선택적 WebView preview"
            req[6] = "사용자가 주요 위치를 외부 지도에서 열 수 있고, 네이티브 지도 SDK 없이도 발표 흐름이 진행된다."
            req[12] = "지도 범위 축소 반영"

        if req_id == "UI-009":
            req[3] = "대표 코스 주변 추천 화면"
            req[4] = (
                "남는 시간에 방문 가능한 대표 코스 주변 카페, 관광지, 포토스팟 후보를 리스트와 위치 확인 버튼으로 보여줘야 한다. "
                "MVP에서는 추천 개수와 지역을 제한한다."
            )
            req[5] = "추천 리스트, 가능/주의 표시, 거리, 외부 지도 링크"
            req[6] = "대표 시나리오에서 남는 시간이 발생하면 가능한 후보가 표시되고, 불가능한 후보는 추천되지 않는다."
            req[7] = "필수"
            req[8] = "상"
            req[9] = "MVP"
            req[12] = "제안서 약속 반영"

        if req_id == "NFR-004":
            req[4] = (
                "백엔드 API가 완성되기 전에도 프론트 화면을 구현하고 발표 흐름을 검증할 수 있도록 고정 JSON 데이터를 제공해야 한다. "
                "단, fixture 데이터는 실제 구현과 다른 fake 계산 경로가 아니라 동일한 DTO와 동일한 계산 엔진 경로를 사용해야 한다."
            )
            req[5] = "fixture JSON, API 응답 계약, 동일 DTO, 동일 엔진, demo mode"
            req[6] = "fixture 데이터만으로 결과 화면이 보이고, 실제 API 전환 시 필드와 계산 경로가 바뀌지 않는다."

        if req_id == "NFR-005":
            req[5] = "API 명세, 요청/응답 JSON, computed/ai 분리, meta 필드, error envelope, 예시 데이터"
            req[6] = "프론트 개발자가 문서만 보고 fixture와 화면을 만들 수 있고, AI 문장과 서버 계산값이 분리되어 있다."

    reqs.extend([
        ["FR-031", "기능", "직접 일정 만들기", "직접 일정 만들기",
         "사용자는 자연어 입력을 사용하지 않고도 날짜, 시작 시간, 출발지, 복귀 지점, 방문 장소, 체류시간을 직접 선택해 일정을 만들 수 있어야 한다.",
         "날짜/시간 선택, 출발지 선택, 복귀지 선택, 장소 검색, 체류시간 입력, 필수 방문 체크",
         "직접 입력한 조건만으로 일정 생성 요청이 만들어지고 자연어 입력 없이도 결과 화면으로 이동할 수 있다.",
         "필수", "상", "MVP", "미시작", "", "사용자 추가 요청"],
        ["FR-032", "기능", "직접 일정 만들기", "장소 추가 및 삭제",
         "사용자는 직접 일정 만들기 과정에서 방문 장소를 추가하거나 삭제할 수 있어야 한다. 삭제된 장소는 일정 계산과 위험도 계산에서 제외되어야 한다.",
         "장소 추가 버튼, 장소 삭제 버튼, 선택 장소 목록, 중복 선택 방지",
         "장소를 추가하면 목록에 반영되고, 삭제하면 일정 생성 요청에서 제외된다.",
         "필수", "상", "MVP", "미시작", "", "사용자 추가 요청"],
        ["FR-033", "기능", "직접 일정 만들기", "장소 순서 직접 조정",
         "사용자는 AI 또는 시스템이 추천한 방문 순서를 그대로 사용하거나, 직접 방문 순서를 바꿀 수 있어야 한다.",
         "장소 순서 변경, 위/아래 이동 또는 드래그 정렬, 순서 저장",
         "장소 순서를 변경하면 변경된 순서 기준으로 이동 구간과 일정 안전도가 다시 계산된다.",
         "중요", "상", "MVP", "미시작", "", "사용자 추가 요청"],
        ["FR-034", "기능", "직접 일정 만들기", "직접 작성 일정 실행 가능성 검사",
         "사용자가 직접 세운 일정도 자연어 입력 일정과 동일하게 버스 시간표, 정류장 후보, 도보 시간, 체류시간을 기준으로 실행 가능성을 검사해야 한다.",
         "직접 작성 일정, TimetableEngine, StopMatcher, RiskEngine, RecoveryEngine",
         "직접 작성 일정에도 안전도 점수, 위험 구간, 출발 권장 시각, 놓침 복구안이 표시된다.",
         "필수", "상", "MVP", "미시작", "", "사용자 추가 요청"],
        ["AI-004", "AI", "직접 일정 검토", "AI 실행 가능성 설명 및 조정 제안",
         "AI는 사용자가 직접 세운 일정에 대해 서버가 계산한 실행 가능성 결과를 바탕으로 왜 가능한지, 왜 위험한지, 어떤 장소나 체류시간을 조정하면 좋은지 설명해야 한다. 단, AI가 버스 시간이나 안전도 점수를 직접 생성해서는 안 된다.",
         "계산 결과 JSON, 위험 사유, 필수/선택 장소, 조정 후보, OpenAI 설명",
         "AI 설명이 서버 계산 결과와 일치하고, 직접 작성 일정의 조정 제안이 계산된 사실을 벗어나지 않는다.",
         "중요", "상", "MVP", "미시작", "", "사용자 추가 요청 + AI guardrail"],
        ["UI-013", "화면", "직접 일정 빌더", "직접 일정 만들기 화면",
         "사용자가 자연어 없이도 여행 조건과 방문 장소를 단계별로 구성할 수 있는 직접 일정 만들기 화면을 제공해야 한다.",
         "단계형 입력 화면, 장소 목록, 체류시간 입력, 필수 방문 체크, 순서 조정, 실행 가능성 검사 버튼",
         "사용자가 직접 일정을 구성하고 실행 가능성 검사 결과로 이동할 수 있다.",
         "필수", "상", "MVP", "미시작", "", "사용자 추가 요청"],
        ["DATA-013", "데이터/API", "데이터 헬스", "데이터 헬스 상태 제공",
         "서비스는 TourAPI cache 상태, 버스 시간표 기준일, fixture/demo mode 여부, fallback 사용 여부, OpenAI fallback 횟수를 확인할 수 있는 데이터 헬스 정보를 제공해야 한다.",
         "GET /api/debug/data-health, dataVersion, cache age, import status, fallback status",
         "데모 또는 운영 중 어떤 데이터 기준으로 계산 중인지 확인할 수 있다.",
         "필수", "상", "MVP", "미시작", "", "Validation/Eng 반영"],
        ["DATA-014", "데이터/API", "Fixture", "fixture와 운영 데이터 동일 경로 사용",
         "동쪽 코리도어 fixture와 실제 import 데이터는 별도 데모 전용 로직이 아니라 같은 테이블, 같은 DTO, 같은 TimetableEngine/RiskEngine/RecoveryEngine을 사용해야 한다.",
         "data_import_runs, fixtureScenarioId, 동일 read model, 동일 계산 엔진",
         "fixture mode와 live mode가 같은 응답 구조를 반환하고, 계산 결과가 다른 코드 경로에서 만들어지지 않는다.",
         "필수", "상", "MVP", "미시작", "", "Eng critical rule"],
        ["UI-011", "화면", "데이터 헬스 화면", "데이터 기준 확인 화면",
         "발표자 또는 운영자는 현재 앱이 실시간 데이터, cache, fixture, fallback 중 어떤 기준으로 동작하는지 확인할 수 있어야 한다.",
         "debug/data-health 화면, 데이터 기준일, TourAPI cache 상태, 시간표 기준일, fallback 표시",
         "데이터 헬스 화면에서 API 장애나 fixture 사용 여부를 확인할 수 있다.",
         "필수", "상", "MVP", "미시작", "", "본선 데모 신뢰 반영"],
        ["UI-012", "화면", "RTO 대시보드", "RTO 미니 대시보드",
         "본선 발표 보조 장면으로 고위험 관광 구간 Top 5, 시간대별 뚜벅이 난이도, 자주 발생하는 대기 위험을 보여주는 미니 대시보드를 제공할 수 있어야 한다.",
         "RTO snapshot fixture, 위험 구간 집계, 시간대별 난이도 카드",
         "사용자 데모 이후 30초 보조 화면으로 지역관광 데이터 가치를 설명할 수 있다.",
         "중요", "중", "Phase 2", "미시작", "", "CEO 계획 반영"],
        ["NFR-006", "비기능", "테스트", "본선 핵심 경로 검수",
         "대표 시나리오는 TourAPI 후보 매칭, 가까운 정류장 후보, 시간표 후보, leave-by, 위험 점수, missed-bus penalty, OpenAI 설명까지 한 번에 검수되어야 한다.",
         "E2E demo script, engine unit test, AI schema/eval fixture, browser/manual checklist",
         "제주공항-함덕-월정리-성산일출봉 입력으로 100/38/81 점수와 +40~48분 위험을 재현한다.",
         "필수", "상", "MVP", "미시작", "", "POC acceptance 반영"],
    ])

    return reqs


requirements = build_requirements()


def border():
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def style_sheet(ws):
    widths = [6, 12, 12, 16, 24, 58, 34, 42, 10, 10, 12, 12, 8, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border()

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border()
        ws.row_dimensions[row[0].row].height = 66

    ws.freeze_panes = "A2"


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "요구사항"

    ws.append(HEADERS)
    for idx, req in enumerate(requirements, 1):
        ws.append([idx] + req)
    style_sheet(ws)

    lists = wb.create_sheet("목록값")
    lists.append(["상태", "체크", "우선순위", "단계"])
    for row in [
        ["미시작", "", "상", "MVP"],
        ["진행중", "✓", "중", "Phase 2"],
        ["완료", "", "하", "Phase 3"],
        ["보류", "", "", ""],
    ]:
        lists.append(row)
    for cell in lists[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border()
    for row in lists.iter_rows(min_row=2):
        for cell in row:
            cell.border = border()
            cell.alignment = Alignment(vertical="center")
    for col in range(1, 5):
        lists.column_dimensions[get_column_letter(col)].width = 16

    summary = wb.create_sheet("요약")
    summary_rows = [
        ["타이밍제주 요구사항 명세서", ""],
        ["작성 기준", "기존 요구사항 명세서 양식 + 2026.06.02 회의 내용"],
        ["총 요구사항", len(requirements)],
        ["기능 요구사항", sum(1 for r in requirements if r[1] == "기능")],
        ["데이터/API 요구사항", sum(1 for r in requirements if r[1] == "데이터/API")],
        ["AI 요구사항", sum(1 for r in requirements if r[1] == "AI")],
        ["화면 요구사항", sum(1 for r in requirements if r[1] == "화면")],
        ["비기능 요구사항", sum(1 for r in requirements if r[1] == "비기능")],
        ["MVP 요구사항", sum(1 for r in requirements if r[9] == "MVP")],
        ["Phase 2 요구사항", sum(1 for r in requirements if r[9] == "Phase 2")],
        ["Phase 3 요구사항", sum(1 for r in requirements if r[9] == "Phase 3")],
        ["핵심 사용자 흐름", "여행 조건 입력 → 관광지 후보 선택 → 버스 시간표 기반 일정 생성 → 일정안 비교 → 출발 권장 시각 확인 → 위험 구간 확인 → 버스 놓침 복구안 확인"],
        ["MVP 범위", "제주공항-함덕-월정리-성산일출봉 동쪽 코스를 우선 완성"],
        ["확장 기능", "실시간 버스, 짐 보관, 날씨 보정, 택시 플랜B, 위젯은 Phase 2 이후로 분리"],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 110
    for row in summary.iter_rows():
        for cell in row:
            cell.border = border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=14)

    wb.active = 0
    return wb


def main():
    wb = build_workbook()
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_REPO)
    print(OUT_REPO.resolve())


if __name__ == "__main__":
    main()
